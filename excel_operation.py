from openpyxl import load_workbook
# read excel file
wb = load_workbook("Book1.xlsx")
ws = wb["Sheet1"]
print(ws.cell(1, 1).value)
# write excel file
ws.cell(1, 1, "India")
wb.save("Book1.xlsx")